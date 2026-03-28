"""
龙虾实验数据自动分析器
- 监控指定文件夹，发现新 Excel 文件自动分析
- 调用 Claude API 生成分析报告
- 报告保存到 reports/ 文件夹
"""

import os
import time
import json
import hashlib
import anthropic
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 配置 ──────────────────────────────────────────────
WATCH_FOLDER = os.getenv("WATCH_FOLDER", "./data")       # 放 Excel 的文件夹
REPORT_FOLDER = os.getenv("REPORT_FOLDER", "./reports")  # 报告输出文件夹
CHECK_INTERVAL = int(os.getenv("CHECK_INTERVAL", "30"))  # 每隔多少秒检查一次
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")   # Railway 变量里配置
# ─────────────────────────────────────────────────────

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

def get_file_hash(filepath):
    """计算文件 hash，用于判断是否已处理过"""
    with open(filepath, "rb") as f:
        return hashlib.md5(f.read()).hexdigest()

def load_processed_log():
    """读取已处理文件记录"""
    log_path = Path(REPORT_FOLDER) / ".processed.json"
    if log_path.exists():
        with open(log_path) as f:
            return json.load(f)
    return {}

def save_processed_log(log):
    """保存已处理文件记录"""
    log_path = Path(REPORT_FOLDER) / ".processed.json"
    with open(log_path, "w") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)

def read_excel(filepath):
    """读取 Excel 文件，返回所有 sheet 的数据摘要"""
    xl = pd.ExcelFile(filepath)
    sheets = {}
    for sheet in xl.sheet_names:
        df = xl.parse(sheet)
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            continue
        sheets[sheet] = {
            "columns": list(df.columns),
            "rows": len(df),
            "preview": df.head(20).to_string(index=False),
            "stats": df.describe(include="all").to_string() if not df.empty else ""
        }
    return sheets

def analyze_with_claude(filename, sheets_data):
    """调用 Claude API 分析数据"""
    content_parts = [f"文件名：{filename}\n"]
    for sheet_name, data in sheets_data.items():
        content_parts.append(f"\n=== Sheet：{sheet_name} ===")
        content_parts.append(f"字段：{', '.join(str(c) for c in data['columns'])}")
        content_parts.append(f"数据行数：{data['rows']}")
        content_parts.append(f"\n数据预览：\n{data['preview']}")
        if data['stats']:
            content_parts.append(f"\n统计摘要：\n{data['stats']}")
    
    user_content = "\n".join(content_parts)

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2000,
        system="""你是一位专业的水产养殖实验数据分析师，专门分析小龙虾养殖实验数据。
请用中文分析实验数据，输出结构如下（直接输出文本，不用 Markdown）：

【数据概述】
简要描述数据内容、实验类型、数据规模。

【关键指标分析】
逐一分析重要指标的数值范围、均值、趋势。

【异常与风险】
指出数据中的异常值、缺失值或值得关注的波动。

【实验结论】
基于数据得出主要结论，哪组/哪个条件表现最好。

【养殖建议】
根据分析结果，给出具体可操作的养殖改进建议。

语言简洁专业，每个部分 3-5 条要点。""",
        messages=[
            {"role": "user", "content": f"请分析以下龙虾实验数据：\n\n{user_content}"}
        ]
    )
    return message.content[0].text

def save_report(filename, analysis_text, sheets_data):
    """将分析结果保存为 Excel 报告"""
    report_name = f"报告_{Path(filename).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    report_path = Path(REPORT_FOLDER) / report_name

    wb = load_workbook() if False else __import__('openpyxl').Workbook()

    # ── 封面 sheet ──
    ws = wb.active
    ws.title = "分析报告"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    header_fill = PatternFill("solid", fgColor="1D6E4F")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=14)
    title_font  = Font(name="微软雅黑", bold=True, size=11)
    body_font   = Font(name="微软雅黑", size=10)
    wrap_align  = Alignment(wrap_text=True, vertical="top")

    # 标题行
    ws.merge_cells("A1:B1")
    ws["A1"] = f"龙虾实验数据分析报告 — {Path(filename).stem}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws["A2"] = "生成时间"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A2"].font = title_font
    ws["B2"].font = body_font

    ws["A3"] = "原始文件"
    ws["B3"] = filename
    ws["A3"].font = title_font
    ws["B3"].font = body_font

    ws.append([])

    # 分析内容逐段写入
    sections = analysis_text.split("【")
    for sec in sections:
        if not sec.strip():
            continue
        if "】" in sec:
            title, content = sec.split("】", 1)
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=f"【{title}】").font = Font(name="微软雅黑", bold=True, size=11, color="1D6E4F")
            ws.cell(row=row, column=2, value=content.strip()).font = body_font
            ws.cell(row=row, column=2).alignment = wrap_align
            ws.row_dimensions[row].height = max(80, content.count("\n") * 16 + 20)
        ws.append([])

    # ── 原始数据 sheet ──
    for sheet_name, data in sheets_data.items():
        try:
            df = pd.read_excel(Path(WATCH_FOLDER) / filename, sheet_name=sheet_name)
            ws2 = wb.create_sheet(title=f"数据_{sheet_name}"[:31])
            # 写表头
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws2.cell(row=1, column=col_idx, value=str(col_name))
                cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2E7D5A")
                cell.alignment = Alignment(horizontal="center")
                ws2.column_dimensions[get_column_letter(col_idx)].width = 16
            # 写数据
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, val in enumerate(row, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=val).font = body_font
        except Exception:
            pass

    wb.save(report_path)
    return report_path

def process_file(filepath, processed_log):
    """处理单个 Excel 文件"""
    filename = Path(filepath).name
    file_hash = get_file_hash(filepath)

    if processed_log.get(str(filepath)) == file_hash:
        return  # 已处理过，跳过

    print(f"[{datetime.now().strftime('%H:%M:%S')}] 发现新文件：{filename}")

    try:
        sheets_data = read_excel(filepath)
        if not sheets_data:
            print(f"  ⚠ 文件为空，跳过")
            return

        print(f"  → 读取成功，共 {len(sheets_data)} 个 sheet，正在分析...")
        analysis = analyze_with_claude(filename, sheets_data)

        report_path = save_report(filename, analysis, sheets_data)
        print(f"  ✓ 报告已生成：{report_path.name}")

        processed_log[str(filepath)] = file_hash
        save_processed_log(processed_log)

    except Exception as e:
        print(f"  ✗ 处理失败：{e}")

def main():
    Path(WATCH_FOLDER).mkdir(parents=True, exist_ok=True)
    Path(REPORT_FOLDER).mkdir(parents=True, exist_ok=True)

    print(f"龙虾AI分析器启动")
    print(f"监控文件夹：{Path(WATCH_FOLDER).resolve()}")
    print(f"报告输出：{Path(REPORT_FOLDER).resolve()}")
    print(f"检查间隔：{CHECK_INTERVAL} 秒\n")

    while True:
        processed_log = load_processed_log()
        excel_files = list(Path(WATCH_FOLDER).glob("*.xlsx")) + \
                      list(Path(WATCH_FOLDER).glob("*.xls"))

        for filepath in sorted(excel_files):
            process_file(filepath, processed_log)

        time.sleep(CHECK_INTERVAL)

if __name__ == "__main__":
    main()
